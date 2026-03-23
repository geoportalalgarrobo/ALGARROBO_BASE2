from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import subprocess
import pathlib
import time
import os

# SOFFICE_PATH: En Railway tratamos de no usarlo para ahorrar espacio
# Pero dejamos la configuración por si se requiere opcionalmente
SOFFICE = os.getenv("SOFFICE_PATH")
if not SOFFICE:
    if os.name == 'nt': # Windows
        SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
    else: # Linux / Docker
        SOFFICE = "/usr/bin/soffice"

def extract_doc(archivo):
    """
    Extracción liviana de archivos .doc (binarios antiguos).
    Prioriza 'antiword' (muy rápido y ligero) sobre 'soffice'.
    """
    archivo = pathlib.Path(archivo)
    
    # 1️⃣ Intentar Antiword (Ideal para entornos muy livianos como Railway)
    try:
        # En el Dockerfile de Railway ya instalamos 'antiword'
        output = subprocess.check_output(["antiword", str(archivo)], stderr=subprocess.STDOUT)
        return output.decode("utf-8", errors="ignore")
    except (subprocess.CalledProcessError, FileNotFoundError):
        # 2️⃣ Fallback a LibreOffice (como estaba originalmente)
        # Solo funcionará si SOFFICE está presente (no es el caso por defecto en Railway)
        if not SOFFICE or not pathlib.Path(SOFFICE).exists():
             return "[Error: No se pudo extraer texto. Instale 'antiword' o 'libreoffice']"
        
        outdir = archivo.parent
        try:
            subprocess.run([
                SOFFICE, "--headless", "--convert-to", "txt:Text",
                str(archivo), "--outdir", str(outdir)
            ], check=True, timeout=10)

            time.sleep(0.5)
            txt_files = list(outdir.glob("*.txt"))
            if not txt_files: return ""
            
            txt_path = max(txt_files, key=lambda p: p.stat().st_mtime)
            content = txt_path.read_text(encoding="utf-8", errors="ignore")
            return content
        except Exception:
            return "[Error al intentar conversión con LibreOffice]"


def extract_text_from_file(file_path, extension):
    try:
        if extension == "pdf":
            reader = PdfReader(file_path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)

        elif extension == "docx":
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)
        
        elif extension == "doc":
            return extract_doc(file_path)

        elif extension in ("xls", "xlsx"):
            wb = load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text.append(str(cell.value))
            return "\n".join(text)

        elif extension in ("png", "jpg", "jpeg"):
            img = Image.open(file_path)
            return pytesseract.image_to_string(img)

        else:
            return ""

    except Exception:
        return ""
